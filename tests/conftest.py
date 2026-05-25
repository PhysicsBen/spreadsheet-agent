"""Fixtures: sample Excel files, test client."""

import io
import json
import zipfile
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient
from langchain_core.messages import AIMessage, HumanMessage, ToolMessage
from openpyxl.worksheet.table import Table

from core.config import settings


@pytest.fixture
def simple_xlsx() -> bytes:
    """Single sheet with simple tabular data, no named tables."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "London"])
    ws.append(["Bob", 25, "Paris"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def named_table_xlsx() -> bytes:
    """Workbook with a named Excel table (ListObject)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Date", "Rep", "Amount"])
    ws.append(["2024-01-01", "Alice", 1000])
    ws.append(["2024-01-02", "Bob", 2000])
    table = Table(displayName="SalesTable", ref="A1:C3")
    ws.add_table(table)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def multi_sheet_xlsx() -> bytes:
    """Workbook with multiple sheets, one hidden."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["ID", "Value"])
    ws1.append([1, 100])

    ws2 = wb.create_sheet("Lookup")
    ws2.append(["Code", "Name"])
    ws2.append(["A", "Alpha"])

    ws3 = wb.create_sheet("Hidden")
    ws3.sheet_state = "hidden"
    ws3.append(["Secret", "Data"])
    ws3.append(["x", "y"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def merged_cells_xlsx() -> bytes:
    """Workbook with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.merge_cells("A1:C1")
    ws["A1"] = "Report Title"
    ws.append(["Name", "Q1", "Q2"])
    ws.append(["Alice", 100, 200])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def formula_xlsx() -> bytes:
    """Workbook with formula cells (no cached values — created without Excel)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = 10
    ws["B1"] = 20
    ws["C1"] = "=A1+B1"  # Formula with no cached value (never calculated by Excel)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def two_region_xlsx() -> bytes:
    """Workbook with two discontiguous table regions separated by two blank rows.

    Two blank rows act as an intentional visual separator between distinct tables.
    A single blank row is treated as intra-table sparsity (see sparse_row_xlsx).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # First table at rows 1-3
    ws.append(["Product", "Price"])
    ws.append(["Apple", 1.5])
    ws.append(["Banana", 0.5])
    # Two blank rows — clear separator between distinct tables
    ws.append([None, None])
    ws.append([None, None])
    # Second table at rows 6-8
    ws.append(["Category", "Count"])
    ws.append(["Fruit", 2])
    ws.append(["Veg", 5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sparse_row_xlsx() -> bytes:
    """Single logical table whose data rows are separated by one blank row.

    Real-world spreadsheets often have sparse rows mid-table (e.g. a missing
    value row, or a blank inserted for readability). The inspector should treat
    a single-blank-row gap as intra-table sparsity and return one table.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "ProjectCode"
    ws.cell(row=1, column=2).value = "Owner"
    ws.cell(row=1, column=3).value = "Budget"
    ws.cell(row=2, column=1).value = "PROJ-A"
    ws.cell(row=2, column=2).value = "Alice"
    ws.cell(row=2, column=3).value = 150000
    ws.cell(row=3, column=1).value = "PROJ-B"  # sparse: only col A
    # row 4 intentionally blank
    ws.cell(row=5, column=1).value = "PROJ-C"
    ws.cell(row=5, column=2).value = "Bob"
    ws.cell(row=5, column=3).value = 80000
    ws.cell(row=6, column=1).value = "PROJ-D"
    ws.cell(row=6, column=2).value = "Carol"
    ws.cell(row=6, column=3).value = 220000
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def annotated_table_xlsx() -> bytes:
    """Sheet with a single-cell annotation row, several blank rows, then a data table.

    The annotation (e.g. a confidentiality notice) should be filtered out as
    noise; only the substantive data table should be returned.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "CONFIDENTIAL — Internal use only"
    # rows 2-4 intentionally blank
    ws.cell(row=5, column=1).value = "Name"
    ws.cell(row=5, column=2).value = "Value"
    ws.cell(row=6, column=1).value = "Alice"
    ws.cell(row=6, column=2).value = 100
    ws.cell(row=7, column=1).value = "Bob"
    ws.cell(row=7, column=2).value = 200
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def formula_xlsx_with_cached() -> bytes:
    """Workbook with formula cells that have cached values (simulates Excel-saved file).

    openpyxl does not compute formulas, so we patch the worksheet XML inside the
    ZIP container to embed a ``<v>`` cached-value element alongside the formula.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CalcCached"
    ws["A1"] = 10
    ws["B1"] = 20
    ws["C1"] = "=A1+B1"  # Formula; no cached value yet

    raw = io.BytesIO()
    wb.save(raw)
    xlsx_bytes = raw.getvalue()

    # Patch the worksheet XML: add <v>30</v> to the C1 formula cell so that
    # openpyxl's data_only=True load returns a non-None cached value.
    buf_out = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zin,
        zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout,
    ):
        for item in zin.infolist():
            data = zin.read(item.filename)
            # Match the worksheet file regardless of capitalisation (openpyxl uses
            # lowercase 'sheet1.xml' but the spec allows other casing).
            if item.filename.lower() == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                # openpyxl writes: <f>A1+B1</f><v /></c> — replace empty value
                # with the cached result so data_only=True returns a non-None value.
                patched = xml.replace(
                    "<f>A1+B1</f><v /></c>", "<f>A1+B1</f><v>30</v></c>"
                )
                # Guard: if the pattern was not found the fixture would be useless.
                assert patched != xml, (
                    "formula_xlsx_with_cached: expected XML pattern not found; "
                    "check openpyxl serialisation format."
                )
                data = patched.encode("utf-8")
            zout.writestr(item, data)
    return buf_out.getvalue()


@pytest.fixture
def legacy_xls_bytes() -> bytes:
    """Minimal legacy .xls file created with xlwt."""
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not installed — skipping legacy .xls fixture")

    wb = xlwt.Workbook()
    ws = wb.add_sheet("LegacySheet")
    ws.write(0, 0, "Product")
    ws.write(0, 1, "Price")
    ws.write(1, 0, "Widget")
    ws.write(1, 1, 9.99)
    ws.write(2, 0, "Gadget")
    ws.write(2, 1, 19.99)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def collision_tables_xlsx() -> bytes:
    """Workbook with two tables sharing the same column names."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collisions"
    ws.append(["Name", "Amount"])
    ws.append(["Alice", 100])
    ws.append(["Bob", 200])
    # Two blank rows — explicit visual separator so these are detected as two
    # distinct tables rather than one sparse table (1 blank row = intra-table
    # sparsity per _ROW_GAP_TOLERANCE; 2+ blank rows = table boundary).
    ws.append([None, None])
    ws.append([None, None])
    ws.append(["Name", "Amount"])
    ws.append(["Carrot", 10])
    ws.append(["Daikon", 20])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class MockCheckpointer:
    """Minimal checkpointer stub used by API tests for thread enumeration/deletion."""

    def __init__(self, parent: "MockGraph") -> None:
        self._parent = parent

    def list(self, _config):
        checkpoints = []
        for thread_id, session_id in self._parent.thread_sessions.items():
            checkpoints.append(
                SimpleNamespace(
                    checkpoint={
                        "channel_values": {
                            "session_id": session_id,
                            "thread_id": thread_id,
                        }
                    },
                    config={"configurable": {"thread_id": thread_id}},
                )
            )
        return checkpoints

    def delete_thread(self, thread_id: str) -> None:
        self._parent.thread_sessions.pop(thread_id, None)
        self._parent.thread_messages.pop(thread_id, None)


class MockGraph:
    """Graph stub used by API tests to avoid real model calls."""

    def __init__(self) -> None:
        self.thread_sessions: dict[str, str] = {}
        self.thread_messages: dict[
            str, list[HumanMessage | AIMessage | ToolMessage]
        ] = {}
        self.checkpointer = MockCheckpointer(self)

    def get_state(self, config):
        thread_id = config["configurable"]["thread_id"]
        messages = self.thread_messages.get(thread_id, [])
        loaded_sheet_names = ["Sheet1"] if messages else []
        return SimpleNamespace(values={"loaded_sheet_names": loaded_sheet_names})

    def invoke(self, initial_state, config):
        thread_id = config["configurable"]["thread_id"]
        session_id = initial_state["session_id"]
        question = initial_state["messages"][-1].content
        prior_messages = self.thread_messages.get(thread_id, [])
        turn_number = len(prior_messages) // 3 + 1

        human = HumanMessage(content=question)
        tool_message = ToolMessage(
            content=json.dumps(
                {
                    "ok": True,
                    "table_id": "Sheet1.table_0",
                    "sheet_name": "Sheet1",
                    "tool_used": "get_sheet_sample",
                }
            ),
            tool_call_id=f"tool-{turn_number}",
            name="get_sheet_sample",
        )
        ai_message = AIMessage(
            content=f"Mock answer turn {turn_number}: {question}",
            response_metadata={
                "model_name": "mock-model",
                "token_usage": {"total_tokens": 21 * turn_number},
            },
            usage_metadata={
                "input_tokens": 10 * turn_number,
                "output_tokens": 11 * turn_number,
                "total_tokens": 21 * turn_number,
            },
        )

        final_messages = [*prior_messages, human, tool_message, ai_message]
        self.thread_sessions[thread_id] = session_id
        self.thread_messages[thread_id] = final_messages

        return {
            **initial_state,
            "messages": final_messages,
            "loaded_sheet_names": ["Sheet1"],
        }


@pytest.fixture
async def api_client(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "db_path", tmp_path / "spreadsheet_agent.db")
    monkeypatch.setattr(settings, "uploads_dir", tmp_path / "uploads")
    monkeypatch.setattr(settings, "query_timeout_secs", 5)
    monkeypatch.setattr(settings, "session_ttl_hours", 24)
    monkeypatch.setattr(settings, "cleanup_interval_hours", 1)

    from api import main as api_main

    mock_graph = MockGraph()
    monkeypatch.setattr(api_main, "graph", mock_graph)

    async with api_main.app.router.lifespan_context(api_main.app):
        transport = ASGITransport(app=api_main.app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            yield client, mock_graph


@pytest.fixture
def random_thread_id() -> str:
    return str(uuid4())
