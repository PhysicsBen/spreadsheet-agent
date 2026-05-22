"""Tests for core/dataframe_loader.py — dual-driver strategy and optional LRU cache."""

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from core.dataframe_loader import clear_cache, get_cache_info, load_sheet, select_engine


@pytest.fixture
def simple_xlsx_file(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "London"])
    ws.append(["Bob", 25, "Paris"])
    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "test.xlsx"
    path.write_bytes(buf.getvalue())
    return path


@pytest.fixture
def multi_sheet_xlsx_file(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["ID", "Value"])
    ws1.append([1, 100])
    ws1.append([2, 200])

    ws2 = wb.create_sheet("Meta")
    ws2.append(["Key", "Val"])
    ws2.append(["version", "1.0"])

    buf = io.BytesIO()
    wb.save(buf)
    path = tmp_path / "multi.xlsx"
    path.write_bytes(buf.getvalue())
    return path


# ── Engine selection ──────────────────────────────────────────────────────────


def test_select_engine_xlsx_returns_calamine(tmp_path):
    p = tmp_path / "file.xlsx"
    p.touch()
    assert select_engine(p) == "calamine"


def test_select_engine_xlsm_returns_calamine(tmp_path):
    p = tmp_path / "file.xlsm"
    p.touch()
    assert select_engine(p) == "calamine"


def test_select_engine_xls_returns_xlrd(tmp_path):
    p = tmp_path / "file.xls"
    p.touch()
    assert select_engine(p) == "xlrd"


def test_select_engine_xlsb_returns_pyxlsb(tmp_path):
    p = tmp_path / "file.xlsb"
    p.touch()
    assert select_engine(p) == "pyxlsb"


# ── Data loading ──────────────────────────────────────────────────────────────


def test_load_sheet_returns_dataframe(simple_xlsx_file):
    df = load_sheet(simple_xlsx_file, "Sheet1")
    assert isinstance(df, pd.DataFrame)
    assert list(df.columns) == ["Name", "Age", "City"]
    assert len(df) == 2


def test_load_sheet_correct_values(simple_xlsx_file):
    df = load_sheet(simple_xlsx_file, "Sheet1")
    assert df.iloc[0]["Name"] == "Alice"
    assert df.iloc[1]["Name"] == "Bob"


def test_load_sheet_with_nrows(simple_xlsx_file):
    df = load_sheet(simple_xlsx_file, "Sheet1", nrows=1)
    assert len(df) == 1
    assert df.iloc[0]["Name"] == "Alice"


def test_load_sheets_returns_dict(multi_sheet_xlsx_file):
    from core.dataframe_loader import load_sheets

    dfs = load_sheets(multi_sheet_xlsx_file, ["Data", "Meta"])
    assert set(dfs.keys()) == {"Data", "Meta"}
    assert isinstance(dfs["Data"], pd.DataFrame)
    assert isinstance(dfs["Meta"], pd.DataFrame)


def test_load_sheets_correct_content(multi_sheet_xlsx_file):
    from core.dataframe_loader import load_sheets

    dfs = load_sheets(multi_sheet_xlsx_file, ["Data"])
    assert list(dfs["Data"].columns) == ["ID", "Value"]
    assert len(dfs["Data"]) == 2


# ── LRU cache ─────────────────────────────────────────────────────────────────


def test_cache_info_accessible():
    info = get_cache_info()
    # Either a CacheInfo namedtuple (when cache enabled) or None (disabled)
    # In tests the default settings have session_cache_size=0 so it may be None
    assert info is None or hasattr(info, "hits")


def test_clear_cache_does_not_raise():
    clear_cache()  # Should be a no-op when cache is disabled
