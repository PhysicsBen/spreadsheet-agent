"""Builds WorkbookMetadataMap at session creation using openpyxl."""

from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

# Number of formula cells to sample when checking for cached values.
_FORMULA_SAMPLE_SIZE = 5

# Maximum rows to scan per sheet when searching for formula cells.
# Pure-data sheets (e.g. CSV-converted xlsx) have no formulas; without this cap
# the read_only pre-scan iterates every cell in every sheet for nothing.
_FORMULA_SCAN_ROW_LIMIT = 1_000

# Sheets with more rows than this skip the full cell scan in heuristic detection.
# Instead, a single table covering the full extent is inferred from the first row.
_LARGE_SHEET_ROW_THRESHOLD = 10_000


def inspect_workbook(
    source: Path | bytes,
    filename: str | None = None,
) -> dict[str, Any]:
    """Build a WorkbookMetadataMap from an Excel file.

    Performs a structural scan of the workbook to capture:
    - Sheet names and visibility state
    - Sheet dimensions (row × col counts)
    - Named Excel tables (ListObjects) — most reliable table boundaries
    - Heuristic contiguous-region tables — fallback when no named tables exist
    - Column names (from table header rows)
    - Merged-cell presence flag per sheet
    - Whether formula cached values are available

    Args:
        source: File path or raw bytes of the Excel file.
        filename: Display filename. Inferred from the path when *source* is a Path.

    Returns:
        WorkbookMetadataMap dict.
    """
    if isinstance(source, Path):
        filename = filename or source.name
    else:
        filename = filename or "unknown.xlsx"

    # Load 1 (fast, read-only): locate formula cells so we can later verify
    # whether cached values exist.  read_only=True avoids loading styles and
    # is substantially faster than a full parse for this sampling step.
    formula_cell_locations = _find_formula_cell_locations(source)

    # Load 2 (full load): structural inspection.  Not read_only so ws.tables
    # and ws.merged_cells are accessible.  data_only=True so cached formula
    # values are returned rather than formula strings.
    if isinstance(source, Path):
        wb = openpyxl.load_workbook(source, data_only=True)
    else:
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True)

    # Reuse the already-loaded workbook to check formula cache — eliminates
    # the third openpyxl parse that the previous implementation required.
    formula_values_available = _check_formula_cache_in_wb(wb, formula_cell_locations)

    sheets_meta: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        state = "visible" if ws.sheet_state == "visible" else "hidden"
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        has_merged = bool(ws.merged_cells.ranges)
        has_errors = _check_error_cells(ws)
        tables = _detect_tables(ws, sheet_name)

        sheets_meta.append(
            {
                "name": sheet_name,
                "state": state,
                "dimensions": {"rows": rows, "cols": cols},
                "has_merged_cells": has_merged,
                "has_error_cells": has_errors,
                "tables": tables,
            }
        )

    wb.close()

    return {
        "filename": filename,
        "sheets": sheets_meta,
        "formula_values_available": formula_values_available,
    }


# ── Internal helpers ──────────────────────────────────────────────────────────


def _find_formula_cell_locations(
    source: Path | bytes,
) -> list[tuple[str, int, int]]:
    """Return up to _FORMULA_SAMPLE_SIZE formula-cell locations as (sheet, row, col).

    Uses read_only=True so openpyxl skips styles and formatting — significantly
    faster than a full load for large files.  data_only=False is required so
    formula strings are returned rather than cached values.
    """
    if isinstance(source, Path):
        wb = openpyxl.load_workbook(source, read_only=True, data_only=False)
    else:
        wb = openpyxl.load_workbook(BytesIO(source), read_only=True, data_only=False)

    formula_cells: list[tuple[str, int, int]] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=_FORMULA_SCAN_ROW_LIMIT):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append((ws.title, cell.row, cell.column))
                    if len(formula_cells) >= _FORMULA_SAMPLE_SIZE:
                        break
            if len(formula_cells) >= _FORMULA_SAMPLE_SIZE:
                break
        if len(formula_cells) >= _FORMULA_SAMPLE_SIZE:
            break

    wb.close()
    return formula_cells


def _check_formula_cache_in_wb(
    wb: openpyxl.Workbook,
    formula_cells: list[tuple[str, int, int]],
) -> bool:
    """Return True if formula cells have cached values, using an already-loaded workbook.

    The workbook must have been loaded with data_only=True.  When no formula
    locations are provided the function returns True (vacuously available).
    """
    if not formula_cells:
        return True

    for sheet_name, target_row, target_col in formula_cells:
        ws = wb[sheet_name]
        cell = ws.cell(row=target_row, column=target_col)
        if cell.value is not None:
            return True

    return False


def _check_error_cells(ws: Worksheet) -> bool:
    """Return True if any cell in the first 100 rows contains an Excel error value."""
    max_row = min(ws.max_row or 0, 100)
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return False

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            # openpyxl represents error cells as objects whose type name contains "Error"
            if cell.value is not None and "Error" in type(cell.value).__name__:
                return True
    return False


def _detect_tables(ws: Worksheet, sheet_name: str) -> list[dict[str, Any]]:
    """Detect tables in a worksheet.

    Precedence:
    1. Named Excel tables (ListObjects via ws.tables) — most reliable.
    2. Heuristic contiguous-region detection — used only when no named tables exist.
    """
    named_tables = _detect_named_tables(ws, sheet_name)
    if named_tables:
        return named_tables
    return _detect_contiguous_regions(ws, sheet_name)


def _detect_named_tables(ws: Worksheet, sheet_name: str) -> list[dict[str, Any]]:
    """Return metadata for all named Excel tables in the worksheet."""
    tables: list[dict[str, Any]] = []

    if not ws.tables:
        return tables

    for tbl in ws.tables.values():
        ref = tbl.ref
        if not ref:
            continue

        min_col, min_row, max_col, _ = range_boundaries(ref)

        columns = []
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row=min_row, column=col_idx)
            val = cell.value
            columns.append(str(val) if val is not None else f"Column{col_idx}")

        display_name = tbl.displayName or tbl.name
        tables.append(
            {
                "id": f"{sheet_name}.{display_name}",
                "type": "named",
                "name": display_name,
                "range": ref,
                "header_row": min_row,
                "columns": columns,
            }
        )

    return tables


def _infer_large_sheet_table(
    ws: Worksheet, sheet_name: str, max_row: int, max_col: int
) -> list[dict[str, Any]]:
    """Return a single inferred table using the first row as column headers.

    Used as a fast fallback when a sheet is too large for the full cell scan.
    """
    columns = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        val = cell.value
        columns.append(str(val) if val is not None else f"Column{c}")

    range_str = f"A1:{get_column_letter(max_col)}{max_row}"
    return [
        {
            "id": f"{sheet_name}.table_0",
            "type": "detected",
            "range": range_str,
            "header_row": 1,
            "columns": columns,
        }
    ]


def _detect_contiguous_regions(ws: Worksheet, sheet_name: str) -> list[dict[str, Any]]:
    """Heuristically detect tables by finding contiguous non-empty cell blocks."""
    max_row = ws.max_row
    max_col = ws.max_column

    if not max_row or not max_col:
        return []

    # For very large sheets, iterating every cell is prohibitively slow.
    # Fall back to inferring a single table from the sheet extent.
    if max_row > _LARGE_SHEET_ROW_THRESHOLD:
        return _infer_large_sheet_table(ws, sheet_name, max_row, max_col)

    # Collect coordinates of non-empty cells
    non_empty: set[tuple[int, int]] = set()
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                non_empty.add((cell.row, cell.column))

    if not non_empty:
        return []

    rows_with_data = sorted({r for r, _ in non_empty})
    row_groups = _find_contiguous_groups(rows_with_data)

    tables: list[dict[str, Any]] = []
    table_idx = 0

    for row_group in row_groups:
        min_row_g = row_group[0]
        max_row_g = row_group[-1]

        cols_in_group = sorted({c for r, c in non_empty if min_row_g <= r <= max_row_g})
        col_groups = _find_contiguous_groups(cols_in_group)

        for col_group in col_groups:
            min_col_g = col_group[0]
            max_col_g = col_group[-1]

            columns = []
            for c in range(min_col_g, max_col_g + 1):
                cell = ws.cell(row=min_row_g, column=c)
                val = cell.value
                columns.append(str(val) if val is not None else f"Column{c}")

            range_str = (
                f"{get_column_letter(min_col_g)}{min_row_g}"
                f":{get_column_letter(max_col_g)}{max_row_g}"
            )

            tables.append(
                {
                    "id": f"{sheet_name}.table_{table_idx}",
                    "type": "detected",
                    "range": range_str,
                    "header_row": min_row_g,
                    "columns": columns,
                }
            )
            table_idx += 1

    return tables


def _find_contiguous_groups(sorted_values: list[int]) -> list[list[int]]:
    """Split a sorted list of integers into groups of consecutive values."""
    if not sorted_values:
        return []

    groups: list[list[int]] = []
    current: list[int] = [sorted_values[0]]

    for v in sorted_values[1:]:
        if v == current[-1] + 1:
            current.append(v)
        else:
            groups.append(current)
            current = [v]

    groups.append(current)
    return groups
